import pandas as pd
from openpyxl import load_workbook, Workbook
from langchain_community.llms import Ollama
from textwrap import dedent
import ollama


llm = Ollama(
    model="nous-hermes2",
    base_url="http://localhost:11434" 
)

question1 = """
        Se requiere escribir una aplicación que se conecte a una base de datos usando JPA y cree una tabla en la base de datos. 
        Luego de ello se solicita explicar: ¿Qué diferencias tiene JPA con JDBC?
        Se evaluará la respuesta del estudiante considerando los siguientes temas: JPA, JDBC. Se espera que en la respuesta del estudiante
        contenga lo siguiente para considerar puntaje completo : persistence.xml, clase entidad JPA, y la clase que cree el entity manager factory.
"""
question2 = """
        Se le da al estudiante el sigueinte caso: Se requiere migrar una applicacion web JSP que trabaja con una version antigua de las librerias de javax.servlet
        a Jakarta version > 9. Ademas, debido a numerosos casos de bugs en desarrollo y pre-producción, la gerencia ha solicitado implementar pruebas unitarias en la aplicación.
        Se solicita al estudiante que de el paso a paso para desarrollar el caso. 
        
        Se evaluará la respuesta del estudiante considerando los siguientes temas: JSP, Servlet, Javax, Jakarta, Tomcat, Wildfly, Junit. Se espera que
        dentro de la respuesta se considere lo siguiente para tener puntaje completo:
        - Se debe revisar si el servidor de aplicaciones soporta la nueva version de Jakarta 
        - Cambiar las dependencias de javax a jakarta en el archivo pom.xml
        - Agregar las dependencias de Junit en el archivo pom.xml
        - Crear las pruebas unitarias.
"""
question3 = """
        Se le solicita al estudiante que desarrolle una aplicación web para una agencia de viajes que permite a los usuarios buscar y reservar paquetes de viaje.
        Se le pide que describa cómo diseñaría e implementaría una funcionalidad que recupere y muestre una lista de paquetes de viaje disponibles desde una base de datos,
        teniendo en cuenta las preferencias del usuario como destino, y fechas de viaje. 

        Se evaluará la respuesta del estudiante considerando los siguientes temas: JPA o JDBC, JSP o JavaServerFaces, Servlets o ManagedBean, Clases para acceso a datos.
"""

class RowData:
    def __init__(self, email, answers):
        self.email = email
        self.answers = answers

class ProcessedData:
    def __init__(self, email, points_q1, points_q2, points_q3):
        self.email = email
        self.points_q1 = points_q1
        self.points_q2 = points_q2
        self.points_q3 = points_q3

class Question:
    def __init__(self, description, score):
        self.description = description
        self.score = score

def extract_data_from_file():
    input_file = 'input.xlsx'
    wb = load_workbook(input_file)
    ws = wb.active
    row_data_list = []
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        email = row[3]  
        answers = [row[4], row[5], row[6]] 
        row_data = RowData(email, answers)
        row_data_list.append(row_data)
    return row_data_list

def write_data_excel(processed_data_list):
    # Create a new workbook and sheet for the output data
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "Processed Data"

    header = ['email', 'points_question1', 'points_question2', 'points_question3']
    ws_output.append(header)

    for data in processed_data_list:
        ws_output.append([data.email, data.points_q1, data.points_q2, data.points_q3])

    output_file = 'output.xlsx'
    wb_output.save(output_file)
    print(f"Processed data has been written to {output_file}")

def process_data () :
    print("Processing data...")
    questions = []
    questions.append(Question(question1, 6))
    questions.append(Question(question2, 6))
    questions.append(Question(question3,8))
    row_data_list = extract_data_from_file()
    data = []
    for row_data in row_data_list:
        conta = 0
        score = []
        for question in questions:   
            sys_prompt=dedent(f"""\
                            Eres un instructor de java senior con experiencia calificando preguntas de examen. 
          
                            Debes calificar la respuesta en base a la pregunta proporcionada, deberas dar una puntuación de 0 a {question.score} puntos. 
                            El resultado final deseado es solamente y UNICAMENTE el puntaje obtenido [INTEGER] solamente ese valor (no extra feedback ni nada, tan solo el puntaje obtenido),
                            por ejemplo si la pregunta en base a tu analisis merece 5 puntos, debes escribir 5 como respuesta.
                            ```
                            Pregunta : 
                                ``` 
                                {question.description}
                                ```
                            Respuesta a calificar : 
                                ```
                                {row_data.answers[conta]}   
                                ```
                            """),
            response = ollama.chat(model='nous-hermes2', messages=[
            {
                'role': 'user',
                'content': str(sys_prompt),
            },
            ])
            if (is_number(response['message']['content'])):
                score.append(response['message']['content'])
            else:
               score.append(-1) # If the response is not a number, we will assign -1, to indicate that the response must be reviewed
            conta += 1

        data.append(ProcessedData(row_data.email, score[0], score[1], score[2]))

    write_data_excel(data)

def is_number(value):
    try:
        float(value)
        return True
    except ValueError:
        return False
                
def main():
    process_data()

if __name__ == "__main__":
    main()












